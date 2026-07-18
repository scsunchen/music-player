import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ==================== Sheet 1: 歌曲表 ====================
ws_songs = wb.active
ws_songs.title = "歌曲"

song_headers = ["id", "title", "artist", "album", "albumId", "duration", "cover", "audioUrl", "mvUrl", "genre"]
song_examples = [
    [1, "晴天", "周杰伦", "十一月的萧邦", 1, 269, "/music-player/images/covers/song1.jpg", "/music-player/audio/song1.mp3", "", "pop"],
    [2, "七里香", "周杰伦", "七里香", 2, 299, "/music-player/images/covers/song2.jpg", "/music-player/audio/song2.mp3", "/music-player/video/mv2.mp4", "pop"],
    [3, "夜曲", "周杰伦", "十一月的萧邦", 1, 225, "/music-player/images/covers/song3.jpg", "/music-player/audio/song3.mp3", "", "pop"],
    [4, "稻香", "周杰伦", "魔杰座", 3, 223, "/music-player/images/covers/song4.jpg", "https://example.com/audio/daoxiang.mp3", "", "pop"],
    [5, "蓝莲花", "许巍", "时光·漫步", 4, 326, "/music-player/images/covers/song5.jpg", "/music-player/audio/song5.mp3", "", "rock"],
]

header_fill = PatternFill("solid", fgColor="2D2D3D")
header_font = Font(bold=True, color="FFFFFF", size=11)
data_font = Font(size=11, color="333333")
zebra_fill_1 = PatternFill("solid", fgColor="FFFFFF")
zebra_fill_2 = PatternFill("solid", fgColor="F7F9FC")
thin_border = Border(
    left=Side("thin", "D9DEE7"), right=Side("thin", "D9DEE7"),
    top=Side("thin", "D9DEE7"), bottom=Side("thin", "D9DEE7")
)

for col_idx, h in enumerate(song_headers, 1):
    cell = ws_songs.cell(1, col_idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, row_data in enumerate(song_examples, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_songs.cell(row_idx, col_idx, val)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = zebra_fill_1 if (row_idx - 2) % 2 == 0 else zebra_fill_2

col_widths_songs = [6, 20, 15, 20, 10, 10, 45, 45, 35, 10]
for i, w in enumerate(col_widths_songs, 1):
    ws_songs.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

# 说明行
note_row = len(song_examples) + 3
ws_songs.cell(note_row, 1, "字段说明:").font = Font(bold=True, color="666666", size=10)
notes = [
    "id - 歌曲唯一ID（数字）",
    "title - 歌曲名",
    "artist - 歌手名",
    "album - 所属专辑名",
    "albumId - 所属专辑ID（对应专辑表的id）",
    "duration - 时长（秒）",
    "cover - 封面图片路径（本地或URL）",
    "audioUrl - 音频文件路径（本地或URL）",
    "mvUrl - MV视频路径（可选，留空表示无MV）",
    "genre - 音乐流派（可选，如 pop/rock/jazz 等）",
]
for i, note in enumerate(notes):
    ws_songs.cell(note_row + 1 + i, 1, note).font = Font(color="888888", size=9)

# ==================== Sheet 2: 专辑表 ====================
ws_albums = wb.create_sheet("专辑")

album_headers = ["id", "name", "artist", "cover", "songs", "releaseDate", "description"]
album_examples = [
    [1, "十一月的萧邦", "周杰伦", "/music-player/images/covers/album1.jpg", "[1, 3]", "2005-11-01", "周杰伦第六张个人专辑"],
    [2, "七里香", "周杰伦", "/music-player/images/covers/album2.jpg", "[2]", "2004-08-03", "周杰伦第五张个人专辑"],
    [3, "魔杰座", "周杰伦", "/music-player/images/covers/album3.jpg", "[4]", "2008-10-15", "周杰伦第九张个人专辑"],
    [4, "时光·漫步", "许巍", "/music-player/images/covers/album4.jpg", "[5]", "2002-12-01", "许巍第二张个人专辑"],
]

for col_idx, h in enumerate(album_headers, 1):
    cell = ws_albums.cell(1, col_idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, row_data in enumerate(album_examples, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_albums.cell(row_idx, col_idx, val)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = zebra_fill_1 if (row_idx - 2) % 2 == 0 else zebra_fill_2

col_widths_albums = [6, 20, 15, 45, 15, 15, 40]
for i, w in enumerate(col_widths_albums, 1):
    ws_albums.column_dimensions[chr(64 + i)].width = w

note_row2 = len(album_examples) + 3
ws_albums.cell(note_row2, 1, "字段说明:").font = Font(bold=True, color="666666", size=10)
album_notes = [
    "id - 专辑唯一ID（数字）",
    "name - 专辑名",
    "artist - 歌手/艺人",
    "cover - 专辑封面路径",
    "songs - 包含的歌曲ID数组（JSON格式，如 [1, 3, 5]）",
    "releaseDate - 发行日期（YYYY-MM-DD）",
    "description - 专辑描述",
]
for i, note in enumerate(album_notes):
    ws_albums.cell(note_row2 + 1 + i, 1, note).font = Font(color="888888", size=9)

# ==================== Sheet 3: 歌单表 ====================
ws_playlists = wb.create_sheet("歌单")

pl_headers = ["id", "name", "description", "cover", "songs", "playCount"]
pl_examples = [
    [1, "周杰伦精选", "精选周杰伦经典歌曲", "/music-player/images/covers/pl1.jpg", "[1, 2, 3, 4]", 125680],
    [2, "摇滚经典", "经典摇滚歌曲合集", "/music-player/images/covers/pl2.jpg", "[5]", 88320],
]

for col_idx, h in enumerate(pl_headers, 1):
    cell = ws_playlists.cell(1, col_idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, row_data in enumerate(pl_examples, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_playlists.cell(row_idx, col_idx, val)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = zebra_fill_1 if (row_idx - 2) % 2 == 0 else zebra_fill_2

col_widths_pl = [6, 20, 30, 45, 20, 12]
for i, w in enumerate(col_widths_pl, 1):
    ws_playlists.column_dimensions[chr(64 + i)].width = w

note_row3 = len(pl_examples) + 3
ws_playlists.cell(note_row3, 1, "字段说明:").font = Font(bold=True, color="666666", size=10)
pl_notes = [
    "id - 歌单唯一ID（数字）",
    "name - 歌单名",
    "description - 歌单描述",
    "cover - 歌单封面路径",
    "songs - 包含的歌曲ID数组（JSON格式，如 [1, 2, 3]）",
    "playCount - 播放量（数字，可选）",
]
for i, note in enumerate(pl_notes):
    ws_playlists.cell(note_row3 + 1 + i, 1, note).font = Font(color="888888", size=9)

# ==================== Sheet 4: 歌词表 ====================
ws_lyrics = wb.create_sheet("歌词")

lyrics_headers = ["songId", "lyrics"]
lyrics_examples = [
    [1, "[00:00.00]晴天\n[00:04.50]故事的小黄花\n[00:09.20]从出生那年就飘着\n[00:13.80]童年的荡秋千\n[00:18.40]随记忆一直晃到现在"],
    [5, "[00:00.00]蓝莲花\n[00:05.20]没有什么能够阻挡\n[00:10.00]你对自由地向往\n[00:15.00]天马行空的生涯\n[00:20.00]你的心了无牵挂"],
]

for col_idx, h in enumerate(lyrics_headers, 1):
    cell = ws_lyrics.cell(1, col_idx, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_idx, row_data in enumerate(lyrics_examples, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_lyrics.cell(row_idx, col_idx, val)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = zebra_fill_1 if (row_idx - 2) % 2 == 0 else zebra_fill_2
        if col_idx == 2:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

ws_lyrics.column_dimensions['A'].width = 10
ws_lyrics.column_dimensions['B'].width = 60

note_row4 = len(lyrics_examples) + 3
ws_lyrics.cell(note_row4, 1, "字段说明:").font = Font(bold=True, color="666666", size=10)
lyrics_notes = [
    "songId - 对应歌曲的ID",
    "lyrics - LRC格式歌词（时间标签 + 歌词文本，换行分隔）",
]
for i, note in enumerate(lyrics_notes):
    ws_lyrics.cell(note_row4 + 1 + i, 1, note).font = Font(color="888888", size=9)

out = "/sessions/6a05e6885ab3f54c2af75600/workspace/music-player/scripts/templates/music_data_template.xlsx"
wb.save(out)
print(f"Excel template saved to {out}")
